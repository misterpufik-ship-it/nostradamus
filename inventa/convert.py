#!/usr/bin/env python3
from __future__ import annotations

import re
import zipfile
from collections import Counter
from io import BytesIO
from xml.etree import ElementTree as ET
from xml.sax.saxutils import escape

from openpyxl import load_workbook

COLUMN_WIDTHS = [
    0.36328125,
    55.6328125,
    22.08984375,
    6.81640625,
    10.453125,
    8.90625,
    15.0,
    0.36328125,
    0.08984375,
    8.0,
    8.5,
    9.90625,
    10.0,
    15.453125,
    11.0,
    6.0,
]
HIDDEN_COLUMNS = {4, 10, 13, 15, 16}
QUANTITY_COLUMN = 11
ORGANIZATION_COLUMN = 14
AUTOFILTER_ORG_COL_ID = 13
MAX_COLUMNS = 16

NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
TAG = lambda name: f"{{{NS}}}{name}"

STYLES_XML = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <fonts count="1">
    <font>
      <sz val="11"/>
      <color indexed="8"/>
      <name val="Aptos Narrow"/>
      <family val="2"/>
      <scheme val="minor"/>
    </font>
  </fonts>
  <fills count="4">
    <fill><patternFill patternType="none"/></fill>
    <fill><patternFill patternType="gray125"/></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FF7DBEFF"/></patternFill></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FFFFFF00"/><bgColor indexed="64"/></patternFill></fill>
  </fills>
  <borders count="1">
    <border><left/><right/><top/><bottom/><diagonal/></border>
  </borders>
  <cellStyleXfs count="1">
    <xf numFmtId="0" fontId="0" fillId="0" borderId="0"/>
  </cellStyleXfs>
  <cellXfs count="3">
    <xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>
    <xf numFmtId="0" fontId="0" fillId="2" borderId="0" xfId="0" applyFill="1"/>
    <xf numFmtId="0" fontId="0" fillId="3" borderId="0" xfId="0" applyFill="1"/>
  </cellXfs>
  <cellStyles count="1">
    <cellStyle name="Normal" xfId="0" builtinId="0"/>
  </cellStyles>
</styleSheet>
"""


def column_index_from_ref(cell_ref: str) -> int:
    letters = re.match(r"^([A-Z]+)", cell_ref)
    if not letters:
        return 0
    value = 0
    for char in letters.group(1):
        value = value * 26 + (ord(char) - 64)
    return value


def detect_primary_organization(source_bytes: bytes, max_row: int) -> str:
    workbook = load_workbook(BytesIO(source_bytes), data_only=True)
    worksheet = workbook.active
    if worksheet is None:
        raise ValueError("В файле нет листов.")

    counts: Counter[str] = Counter()
    for row_number in range(2, max_row + 1):
        value = worksheet.cell(row_number, ORGANIZATION_COLUMN).value
        if value in (None, ""):
            continue
        counts[str(value).strip()] += 1

    if not counts:
        raise ValueError("Не удалось определить организацию в колонке N.")
    return counts.most_common(1)[0][0]


def build_cols_xml() -> str:
    parts = ["<cols>"]
    for index, width in enumerate(COLUMN_WIDTHS, start=1):
        attrs = [
            f'min="{index}"',
            f'max="{index}"',
            f'width="{width}"',
            'customWidth="1"',
        ]
        if index in HIDDEN_COLUMNS:
            attrs.append('hidden="1"')
        if index == QUANTITY_COLUMN:
            attrs.append('style="2"')
        parts.append(f"<col {' '.join(attrs)}/>")
    parts.append("</cols>")
    return "".join(parts)


def build_autofilter_xml(max_row: int, organization: str) -> str:
    return (
        f'<autoFilter ref="A1:P{max_row}">'
        f'<filterColumn colId="{AUTOFILTER_ORG_COL_ID}">'
        f"<filters><filter val=\"{escape(organization)}\"/></filters>"
        "</filterColumn></autoFilter>"
    )


def patch_sheet_xml(sheet_xml: bytes, max_row: int, organization: str) -> bytes:
    text = sheet_xml.decode("utf-8")

    text = re.sub(
        r"<dimension\b[^>]*/>",
        f'<dimension ref="A1:P{max_row}"/>',
        text,
        count=1,
    )

    cols_xml = build_cols_xml()
    if "<cols>" in text:
        text = re.sub(r"<cols>.*?</cols>", cols_xml, text, count=1, flags=re.S)
    else:
        text = re.sub(
            r"(<sheetFormatPr\b[^>]*/>)",
            r"\1" + cols_xml,
            text,
            count=1,
        )

    autofilter_xml = build_autofilter_xml(max_row, organization)
    text = re.sub(r"<autoFilter\b[\s\S]*?</autoFilter>", "", text)
    text = re.sub(r"(</cols>)", r"\1" + autofilter_xml, text, count=1)

    root = ET.fromstring(text.encode("utf-8"))
    for row in root.findall(f".//{TAG('row')}"):
        row_number = int(row.get("r", "0"))
        for cell in row.findall(TAG("c")):
            ref = cell.get("r", "")
            column_number = column_index_from_ref(ref)
            if column_number < 1 or column_number > MAX_COLUMNS:
                continue

            if row_number == 1:
                style_id = "2" if column_number == QUANTITY_COLUMN else "1"
                cell.set("s", style_id)
            elif column_number == QUANTITY_COLUMN:
                cell.set("s", "2")
            elif "s" in cell.attrib:
                del cell.attrib["s"]

    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def patch_workbook_xml(workbook_xml: bytes, sheet_name: str, max_row: int) -> bytes:
    text = workbook_xml.decode("utf-8")
    escaped_sheet = sheet_name.replace("'", "''")
    defined_name = (
        f"<definedName name=\"_xlnm._FilterDatabase\" localSheetId=\"0\" hidden=\"1\">"
        f"'{escaped_sheet}'!$A$1:$P${max_row}</definedName>"
    )

    if "_xlnm._FilterDatabase" in text:
        text = re.sub(
            r"<definedName name=\"_xlnm\._FilterDatabase\"[\s\S]*?</definedName>",
            defined_name,
            text,
            count=1,
        )
    else:
        text = text.replace(
            "</workbook>",
            f"<definedNames>{defined_name}</definedNames></workbook>",
        )

    return text.encode("utf-8")


def get_sheet_name(workbook_xml: bytes) -> str:
    root = ET.fromstring(workbook_xml)
    sheet = root.find(f".//{TAG('sheet')}")
    if sheet is None or not sheet.get("name"):
        return "Sheet1"
    return sheet.get("name", "Sheet1")


def get_max_row(sheet_xml: bytes) -> int:
    root = ET.fromstring(sheet_xml)
    max_row = 1
    for row in root.findall(f".//{TAG('row')}"):
        max_row = max(max_row, int(row.get("r", "1")))
    return max_row


def convert_bytes(source_bytes: bytes) -> tuple[bytes, dict]:
    with zipfile.ZipFile(BytesIO(source_bytes), "r") as source_zip:
        sheet_name = "sheet1.xml"
        for name in source_zip.namelist():
            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
                sheet_name = name.split("/")[-1]
                break

        sheet_path = f"xl/worksheets/{sheet_name}"
        sheet_xml = source_zip.read(sheet_path)
        workbook_xml = source_zip.read("xl/workbook.xml")
        max_row = get_max_row(sheet_xml)
        organization = detect_primary_organization(source_bytes, max_row)
        sheet_name_label = get_sheet_name(workbook_xml)

        patched_sheet = patch_sheet_xml(sheet_xml, max_row, organization)
        patched_workbook = patch_workbook_xml(workbook_xml, sheet_name_label, max_row)
        styles_xml = STYLES_XML.encode("utf-8")

        output = BytesIO()
        with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as target_zip:
            for item in source_zip.infolist():
                data = source_zip.read(item.filename)
                if item.filename == sheet_path:
                    data = patched_sheet
                elif item.filename == "xl/workbook.xml":
                    data = patched_workbook
                elif item.filename == "xl/styles.xml":
                    data = styles_xml
                target_zip.writestr(item, data)

    meta = {
        "rows": max_row,
        "sheetName": sheet_name_label,
        "organization": organization,
    }
    return output.getvalue(), meta
